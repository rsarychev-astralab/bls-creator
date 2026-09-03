from datetime import datetime
from pathlib import Path
from re import sub

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import ROOT
from app.output.gas import gas_payload

RESULTS_DIR = ROOT / "data" / "results"

BLUE = PatternFill("solid", fgColor="4A86E8")
GREEN = PatternFill("solid", fgColor="D9EAD3")
HEAD = PatternFill("solid", fgColor="C9DAF8")
WHITE = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADERS = [
    "Вариант ответа",
    "Контакт",
    "Не контакт",
    "Контакт %",
    "Неконтакт %",
    "Прирост",
    "Направление",
]


def _paint(ws, row: int, fill: PatternFill, font: Font | None = None) -> None:
    for col in range(1, 8):
        cell = ws.cell(row, col)
        cell.fill = fill
        if font:
            cell.font = font


def _render_question(ws, question: dict, q_num: int, start: int, totals: dict) -> int:
    row = start
    metric = (question.get("metric") or "").strip()
    if metric:
        ws.cell(row, 1, metric).font = BOLD
        _paint(ws, row, GREEN, BOLD)
        row += 1
    ws.cell(row, 1, f"Вопрос {q_num}:").font = BOLD
    ws.cell(row, 2, question.get("question_text") or "").font = BOLD
    _paint(ws, row, GREEN, BOLD)
    row += 1

    responses = question.get("responses") or []
    if not responses:
        return row

    header_row = row
    for i, title in enumerate(HEADERS, start=1):
        ws.cell(row, i, title).font = BOLD
    _paint(ws, row, HEAD, BOLD)
    row += 1
    data_start = row
    n = len(responses)
    contact = int((totals or {}).get("total_contact_group") or 400)
    noncontact = int((totals or {}).get("total_noncontact_group") or 400)

    for resp in responses:
        ws.cell(row, 1, resp.get("option") or "")
        ws.cell(row, 2, int(resp.get("contact_count") or 0))
        ws.cell(row, 3, int(resp.get("noncontact_count") or 0))
        row += 1

    row = data_start + n + 1
    ws.cell(row, 1, "Итого:").font = BOLD
    row += 1
    ws.cell(row, 1, "Всего ответов:").font = BOLD
    ws.cell(row, 2, f"=SUM(B{data_start}:B{data_start + n - 1})")
    ws.cell(row, 3, f"=SUM(C{data_start}:C{data_start + n - 1})")
    row += 1
    respondents_row = row
    ws.cell(row, 1, "Всего респондентов:").font = BOLD
    ws.cell(row, 2, contact)
    ws.cell(row, 3, noncontact)
    row += 1
    ws.cell(row, 1, "Прирост составил").font = BOLD
    ws.cell(row, 2, f"=MAX(F{data_start}:F{data_start + n - 1})")
    ws.cell(row, 2).number_format = "+0.00%;-0.00%;0.00%"
    row += 1

    for i in range(n):
        r = data_start + i
        ws.cell(r, 4, f"=B{r}/$B${respondents_row}")
        ws.cell(r, 5, f"=C{r}/$C${respondents_row}")
        ws.cell(r, 6, f"=D{r}-E{r}")
        ws.cell(r, 7, f'=IF(F{r}>0,"Положительный",IF(F{r}<0,"Отрицательный","Без прироста"))')
        ws.cell(r, 4).number_format = "0.00%"
        ws.cell(r, 5).number_format = "0.00%"
        ws.cell(r, 6).number_format = "+0.00%;-0.00%;0.00%"
        for col in range(1, 8):
            ws.cell(r, col).border = THIN
            ws.cell(header_row, col).border = THIN

    green = FormulaRule(
        formula=[f"$F{data_start}>0"],
        fill=PatternFill("solid", fgColor="B6D7A8"),
        font=Font(color="274E13"),
    )
    red = FormulaRule(
        formula=[f"$F{data_start}<0"],
        fill=PatternFill("solid", fgColor="EA9999"),
        font=Font(color="660000"),
    )
    rng = f"G{data_start}:G{data_start + n - 1}"
    ws.conditional_formatting.add(rng, green)
    ws.conditional_formatting.add(rng, red)
    return row


def write_result_xlsx(modeled: object) -> dict:
    payload = gas_payload(modeled)
    first = payload[0] if payload else {}
    project_name = next(iter(first.keys()), "Brand Lift Report") if isinstance(first, dict) else "Brand Lift Report"
    clean = sub(r'[\\/:*?"<>|]', "_", str(project_name))[:80]
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    title = f"{clean} — {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    filename = f"{clean}_{stamp}.xlsx"
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    path = RESULTS_DIR / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "BLS"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 56
    for col, width in enumerate([10, 10, 12, 10, 16], start=3):
        ws.column_dimensions[get_column_letter(col)].width = width

    row = 1
    for idx, project_obj in enumerate(payload):
        if not isinstance(project_obj, dict) or not project_obj:
            continue
        key = next(iter(project_obj.keys()))
        project = project_obj.get(key) or {}
        ws.cell(row, 1, f"Проект {idx + 1}:").font = WHITE
        ws.cell(row, 2, key).font = WHITE
        _paint(ws, row, BLUE, WHITE)
        row += 1
        meta = project.get("metadata") or {}
        totals = {
            "total_contact_group": meta.get("total_contact_group") or 400,
            "total_noncontact_group": meta.get("total_noncontact_group") or 400,
        }
        for q_idx, question in enumerate(project.get("questions") or [], start=1):
            row = _render_question(ws, question, q_idx, row, totals)
            row += 2

    wb.save(path)
    return {
        "spreadsheet_name": title,
        "filename": filename,
        "path": str(path),
        "spreadsheet_url": f"/api/results/{filename}",
        "spreadsheet_id": "",
    }
